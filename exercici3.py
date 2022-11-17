
#Dades a introduir a csv i excel
data = [["Name", "Surname", "Age"], ["Roger","Sobrino","45"]]

#Creem la funcio que escriu a csv i despres el csv a consola
def csv_consola():
    #Importem el paquet csv
    import csv
    #Es crea un arxiu anomenat "roger.csv", el mode 'w' es d'escriptura, i aixo ho inicialitzem com a variable "arxiu"
    with open('roger.csv', mode='w') as arxiu:
        #Creem una variable que escriura a "arxiu. El delimitador son les cometes (name es un valor, surname un altre...)
        #Quotechar indica quin caracter es fara servir per denotar camps (en aquest cas "), i quoting indica, en aquest cas,
        #que nomes s'escriuran els valors que estiguin entre " (indicat a quotechar)
        roger_writer = csv.writer(arxiu, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        #S'escriuran els valors de data, desde el [0] fins a l'ultim
        roger_writer.writerows(data[0:])

    #S'obre l'arxiu "roger.csv", i guardem aixo com a variable arxiu
    with open('roger.csv') as arxiu:
        #La variable roger_reader llegira l'arxiu obert, i tindra en compte els camps delimitats per comes
        roger_reader = csv.reader(arxiu, delimiter=',')
        #Per cada valor d'una fila a l'arxiu...
        for row in roger_reader:
            #S'aniran imprimint amb un espai entre ells
            print(" ".join(row))

#Creem la funcio que creara i llegira excel
def excel():
    #Importem el paquet openpyxl
    import openpyxl
    #Fem una variable que sera un full de calcul
    wb = openpyxl.Workbook()
    #la variable full sera el full que estigui actiu
    full = wb.active
    #c1 es la cel·la de la fila 1 i columna 1...
    c1 = full.cell(row=1, column=1)
    #... i li assignem el primer valor de data
    c1.value=data[0][0]
    #Passa el mateix amb la resta de cel·les, les anem omplint de valors.
    c2 = full.cell(row=1, column=2)
    c2.value=data[0][1]
    c3 = full.cell(row=1, column=3)
    c3.value = data[0][2]
    c4 = full.cell(row=2, column=1)
    c4.value = data[1][0]
    c5 = full.cell(row=2, column=2)
    c5.value = data[1][1]
    c6 = full.cell(row=2, column=3)
    c6.value = data[1][2]
    #Guardem el full en un arxiu, que es troba a la rula seguent
    wb.save(".\\resultat.xlsx")

    #Ara farem la part que llegeix l'excel
    #Indiquem el path (es el mateix que hem creat)
    path=".\\resultat.xlsx"
    #Aquesta variable carrega el full que es troba al path indicat
    wb_obj=openpyxl.load_workbook(path)
    #Aquesta variable indica el full actiu
    sheet_obj = wb_obj.active
    #Imprimim els valors de cadascuna de les cel·les indicades del full actiu
    print(f'{sheet_obj["A1"].value}: {sheet_obj["A2"].value}')
    print(f'{sheet_obj["B1"].value}: {sheet_obj["B2"].value}')
    print(f'{sheet_obj["C1"].value}: {sheet_obj["C2"].value}')


#Invoquem les funcions que hem fet
csv_consola()
excel()